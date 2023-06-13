import json
import pandas
import os
import openpyxl
import ctypes 
import argparse
import numpy
import matplotlib
import matplotlib.pyplot
import scipy
import scipy.stats
import platform


# 判断是否为路径
def is_dir(file_path):
    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)

# 指令输入及传出
def cmd_inter():
    # 创建一个ArgumentParser的对象，此对象包含将命令行参数解析为 python 数据类型所需的全部信息。
    parser = argparse.ArgumentParser(description='Generates the json of google benchmark to csv/excel/polt.', add_help=True)
    # 添加参数命令
    parser.add_argument('-i', '--input', help="Input .json files directory.")
    parser.add_argument('-o', '--output', help="Output .csv/.xlsx/.png files directory.")
    parser.add_argument('-l', '--link', help="Link c/c++ libraries directory.(not support)")
    parser.add_argument('-m', '--mode', help="Output all/simple/user-defined metrics content to .csv/.xlsx/.png.", default=["simple"], nargs="*")
    parser.add_argument('-s', '--select', help="Select the output files from csv/xlsx/png.", default=["csv", "xlsx", "png"], nargs="*")
    parser.add_argument('-n', '--name', help="Need analysis .json files name to excel.", default=["all"], nargs="*")
    parser.add_argument('-p', '--plot', help="From all .json files, plot term only suport cpu_time/real_time now.", default=["cpu_time", "real_time"], nargs="*")
    parser.add_argument('-f', '--apip', help="Point api with different parameters to get .png files.", default=["yes"])

    # 初步解析参数命令
    args = parser.parse_args()
    if args.input is None:
        args.input = os.getcwd()

    if args.output is None:
        args.output = os.getcwd()
        is_dir(args.output)
    
    if args.link is None:
        args.link = os.path.dirname(__file__)
    print(args)
    return args

# 扫描路径下的所有文件类型
def scan_filetype(directory, filetype):
    files = []
    for entry in os.scandir(directory):
        if entry.path.endswith(filetype) and entry.is_file():
            files.append(entry)
    if len(files) == 0:
        print('Error no such file type! ')
    return files, entry

# 读取json文件所有数据
def read_json(filename: str) -> dict:
    try:
        with open(filename, "r") as f:
            data = json.loads(f.read())
    except:
        raise Exception(f"Reading {filename} file encountered an error")
    return data

# 解析单个json文件数据
def analysis_file(file, arga):
    print('analysis ' + file)
    data = read_json(file)
    
    # dict to list.
    context = []
    if type(data).__name__=='dict':
        if 'context' in data.keys():
            for a in data['context']:
                context_dict = {}
                if a != "caches":
                    context_dict['name'] = a
                    context_dict['parameter'] =  data['context'][a]
                    context.append(context_dict)

    context_data = pandas.json_normalize(context)
    # print(context_data)

    # cpu caches per cpu.
    if type(data).__name__=='dict':
        if 'context' in data.keys() and 'caches' in data['context'].keys(): 
            type_data = []; level_data = []; size_data = []; num_sharing_data = []
            for b in data['context']['caches']:
                type_data.append(b['type'])
                level_data.append(b['level'])
                size_data.append(b['size'])
                num_sharing_data.append(b['num_sharing'])
            caches = [
                {
                    "name": 'type',
                    "parameter": type_data
                },
                {
                    "name": 'level',
                    "parameter": level_data
                },
                {
                    "name": 'size',
                    "parameter": size_data
                },
                {
                    "name": 'num_sharing',
                    "parameter": num_sharing_data
                }
            ]
            caches_data = pandas.json_normalize(caches)
            # print(caches_data)

    # system_software_hardware info
    system_sh = [
        # {
        #     "name": 'host_name',
        #     " parameter": platform.node()
        # },
        {
            "name": 'cpu',
            "parameter": platform.processor()
        },
        {
            "name": 'kernel',
            "parameter": platform.release()
        },
        {
            "name": "compiler",
            "parameter": platform.python_compiler()
        },
        {
            "name": 'arch',
            "parameter": platform.architecture()
        }
    ]
    system_sh_data = pandas.json_normalize(system_sh)
    # print(system_sh_data)

    # hardware info
    hardware = [
        {
            "name": 'device_id', 
            "parameter": 0
        },
        {
            "name": 'baord_name', 
            "parameter": 'M3'
        },
        {
            "name": 'SN',
            "parameter": 12345678
        }
    ]
    hardware_data = pandas.json_normalize(hardware)
    # print(hardware_data)

    # link c/c++ lib .so/.a
    perfinfo = ctypes.cdll.LoadLibrary(os.path.dirname(__file__) + "/libperfinfo.so")

    # c string format transfer to python string
    perfinfo.getPlatform.restype = ctypes.c_char_p
    perfinfo.getMpulib.restype = ctypes.c_char_p
    perfinfo.getMps.restype = ctypes.c_char_p
    perfinfo.getDaemon.restype = ctypes.c_char_p

    # config info
    config = [
        {
            "name": 'platform',
            "parameter": perfinfo.getPlatform().decode("utf-8")
        },
        {
            "name": 'mpulib',
            "parameter": perfinfo.getMpulib().decode("utf-8")
        },
        {
            "name": 'mps',
            "parameter": perfinfo.getMps().decode("utf-8")
        },
        {
            "name": 'daemon',
            "parameter": perfinfo.getDaemon().decode("utf-8")
        },
        {
            "name": 'loglevel',
            "parameter": perfinfo.getLoglevel()
        }
    ]
    config_data = pandas.json_normalize(config)
    # print(config_data)
    
    # filt key for dict.
    if type(data).__name__=='dict':
        if 'benchmarks' in data.keys():
            benchmark = data['benchmarks']
            for c in benchmark:
                time_unit = c['time_unit']
                c['cpu_time' + '(' + time_unit + ')'] = c.pop('cpu_time')
                c['real_time' + '(' + time_unit + ')'] = c.pop('real_time')
            benchmark_data = pandas.json_normalize(benchmark)
            # print(benchmark_data)
            host_data = pandas.concat([context_data, system_sh_data, caches_data], axis = 0)
            host_data.reset_index(drop=True, inplace=True)
            # print(host_data)
            perf_env_data = pandas.concat([hardware_data, host_data, config_data], keys = ['hardware_info', 'host_info', 'config_info'], names=['Class','Index'], axis = 0)
            print(perf_env_data)
            if 'simple' in arga:
                benchmark_filt = ['name', 'family_index', 'real_time' + '(' + time_unit + ')', 'cpu_time' + '(' + time_unit + ')', 'iterations']
                print(benchmark_data[benchmark_filt])
                return benchmark_data[benchmark_filt], perf_env_data
            elif 'all' in arga:
                print(benchmark_data)
                return benchmark_data, perf_env_data
            else:
                if 'cpu_time' in arga:
                    arga[arga.index('cpu_time')] = 'cpu_time' + '(' + time_unit + ')'
                if 'real_time' in arga:
                    arga[arga.index('real_time')] = 'real_time' + '(' + time_unit + ')'
                print(benchmark_data[arga])
                return benchmark_data[arga], perf_env_data
    return None, None

# 解析json文件数据
def analysis_files(file, plots_data, arga):
    print('analysis ' + file)
    data = read_json(file)

    # # dict to list.
    # context = []
    # for a in data['context']:
    #     context_dict = {}
    #     if a != "caches":
    #         context_dict['name'] = a
    #         context_dict['parameter'] =  data['context'][a]
    #         context.append(context_dict)

    # context_data = pandas.json_normalize(context)
    # print(context_data)

    # # cpu caches per cpu.
    # caches = data['context']['caches']
    # caches_data = pandas.json_normalize(caches)
    # print(caches_data)

    for b in data['benchmarks']:
        # print('\t \033[0;34m%-16s\033[0m ----> \033[0;32m%-64s\033[0m = \033[0;33m%-32s\033[0m'%(arga, b['name'], str(b[arga])))
        if plots_data.get(b['name']) is None:
            plots_data[b['name']] = [b[arga]]
        else:
            plots_data[b['name']].append(b[arga])

# 针对统一api不同参数的解析
def  api_analysis_files(file, plots_data, name, arga):
    print('analysis ' + file)
    data = read_json(file)
    for a in data['benchmarks']:
        if plots_data.get(a['name'].split('/')[1]) is None:
            if a['name'].split('/')[1] not in name:
                name.append(a['name'].split('/')[1])
            plots_data[a['name'].split('/')[1]] = [a[arga]]
            if len(a['name'].split('/')) > 2:
                plots_data[a['name'].split('/')[1] + '/p'] = [a['name'].replace(a['name'].split('/')[0]+ '/' + a['name'].split('/')[1] + '/', '')]
        else:
            plots_data[a['name'].split('/')[1]].append(a[arga])
            if len(a['name'].split('/')) > 2:
                plots_data[a['name'].split('/')[1] + '/p'].append(a['name'].replace(a['name'].split('/')[0]+ '/' + a['name'].split('/')[1] + '/', ''))


def smooth(x,window_len=11,window='hanning'):
    # references: https://scipy-cookbook.readthedocs.io/items/SignalSmooth.html
    if x.ndim != 1:
        raise ValueError("smooth only accepts 1 dimension arrays.")

    if x.size < window_len:
        raise ValueError("Input vector needs to be bigger than window size.")

    if window_len<3:
        return x

    if not window in ['flat', 'hanning', 'hamming', 'bartlett', 'blackman']:
        raise ValueError("Window is on of 'flat', 'hanning', 'hamming', 'bartlett', 'blackman'")

    s=numpy.r_[x[window_len-1:0:-1],x,x[-2:-window_len-1:-1]]
    if window == 'flat': #moving average
        w=numpy.ones(window_len,'d')
    else:
        w=eval('numpy.'+window+'(window_len)')

    y=numpy.convolve(w/w.sum(),s,mode='valid')
    return y[0:len(x)]  

def hasSlowedDown(benchmark, raw_values, smoothedvalues, slidingwindow, alphavalue, metric):
    sample_count = len(raw_values)
    sample_a_len = sample_count - slidingwindow
    sample_b_len = slidingwindow

    # mw test
    sample_a = smoothedvalues[:sample_a_len]
    sample_b = smoothedvalues[sample_a_len:]
    print('len(sample_a) = ' + str(len(sample_a)) + ' len(sample_b) = ' + str(len(sample_b)))
    stat, p = scipy.stats.mannwhitneyu(sample_a, sample_b)
    print('BENCHMARK ' + benchmark + ' STATS=%.3f, p=%.3f' % (stat, p))
    if p < alphavalue:
        print('\tStep change possibly found, performing t-test...')
        
        # confirm with Welch's t-test as mw can reject if sd is big (see: https://thestatsgeek.com/2014/04/12/is-the-wilcoxon-mann-whitney-test-a-good-non-parametric-alternative-to-the-t-test/)
        stat, p = scipy.stats.ttest_ind(sample_a, sample_b, equal_var = False)
        if p < alphavalue:
            return True;
        print('\tStep change doesnt appear to be part of a trend')
    return False 

def turningpoints(x):
    peaks = []
    troughs = []
    for i in range(1, len(x)-1):
        if (x[i-1] < x[i] and x[i+1] < x[i]):
            peaks.append(i)
        elif (x[i-1] > x[i] and x[i+1] > x[i]):
            troughs.append(i)
    return peaks, troughs


def estimateStepLocation(values):
    # references: https://stackoverflow.com/questions/48000663/step-detection-in-one-dimensional-data/48001937)
    dary = numpy.array(values)
    avg = numpy.average(dary)
    dary -= avg
    step = numpy.hstack((numpy.ones(len(dary)), -1*numpy.ones(len(dary))))
    dary_step = numpy.convolve(dary, step, mode='valid')
    print(numpy.argmax(dary_step))

    # get location of step change
    peaks, troughs = turningpoints(dary_step)
    if(len(peaks)) == 0:
        return 0;

    step_max_idx = peaks[-1]
    return step_max_idx

def auto_width(filename):
    # 打开xlsx文件
    wb = openpyxl.load_workbook(filename)
    # 遍历sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pandas.read_excel(filename,sheet_name=sheet, engine='openpyxl')
        # 把表头改到最后一行
        df.loc[len(df)]=list(df.columns)
        for i in df.index:
            # 设置单元格对齐方式 Alignment(horizontal=水平对齐模式,vertical=垂直对齐模式,text_rotation=旋转角度,wrap_text=是否自动换行)
            alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
            ws['A'+str(i+1)].alignment = alignment
            ws['B'+str(i+1)].alignment = alignment
            ws.row_dimensions[i].height = 20
        for col in df.columns:
            # 获取列序号
            index = list(df.columns).index(col)
            # 获取行字母表头
            letter = openpyxl.utils.get_column_letter(index+1)
            # 获取当前列最大宽度
            collen = df[col].apply(lambda x :len(str(x).encode())).max()
            # 设置列宽为最大长度比例
            ws.column_dimensions[letter].width = collen * 1.2
    wb.save(filename)


def name_cut(name):
    return name.replace(name.split('/')[0] + '/', '')


def plot_2d(x_data, y_data, arga, output, name):
    fig = matplotlib.pyplot.figure(figsize=(16, 9))
    ax = fig.add_subplot()
    if len(y_data) > 1: 
        # get the same test items data 
        y_raw_val = y_data
        x_raw_val  = x_data
        if x_raw_val is None:
            x_raw_val = range(0, len(numpy.arange(0, len(y_raw_val), 1)))

        matplotlib.pyplot.plot(numpy.arange(0, len(y_raw_val), 1), y_raw_val, color='darkgray', label="raw_linear")
        matplotlib.pyplot.scatter(numpy.arange(0, len(y_raw_val), 1), y_raw_val, color='black', label="raw_values", s=12)
        
        if len(y_raw_val) > 8:
            # 1D data smooth
            smoothed_val = smooth(numpy.array(y_raw_val), 8)
            matplotlib.pyplot.plot(smoothed_val, '-b', label="smoothed_values")

        # generate x data step = 1, min = 0, max = strlen(raw_val)
        x_vals  = numpy.arange(0, len(y_raw_val), 1)
        y_vals  = y_raw_val
        # Least Squares Fitting
        model   = numpy.polyfit(x_vals, y_vals, 1)
        predict = numpy.poly1d(model)
        lrx     = range(0, len(x_vals))
        lry     = predict(lrx)
        matplotlib.pyplot.plot(lry, 'tab:orange', label="fit_linear")

        #
        matplotlib.pyplot.xticks(numpy.arange(0, len(y_raw_val)), labels=x_raw_val, color='navy', rotation=45)
        ax.xaxis.set_major_locator(matplotlib.ticker.MultipleLocator(len(y_raw_val)/40))

        # figure files name
        matplotlib.pyplot.ylabel(arga)
        matplotlib.pyplot.xlabel('sample#')
        # print(a.split('/')[0])
        matplotlib.pyplot.title(name, color='blue', fontstyle='italic', loc ='right')
        figurename = '{}-{}.png'.format(arga, name.replace('/', '-').replace(':', ''))
        matplotlib.pyplot.legend()
        matplotlib.pyplot.tight_layout()
        figurepath = os.path.join(output, 'figure/' + figurename)
        is_dir(figurepath)
        matplotlib.pyplot.savefig(figurepath)
        matplotlib.pyplot.clf()
        matplotlib.pyplot.close() 

def main():
    # cmd
    args = cmd_inter()

    # get list of files to parse
    files, entry = scan_filetype(args.input, ".json")

    # get data from all json to plot.
    plots = args.plot
    benchmarks_plots = {}
    api_plots = {}
    api_names = []
    for plot in  plots:
        plots_data = {}
        api_data = {}
        for entry in files:
           if entry.path.endswith('.json') and entry.is_file(): 
                try:
                    if "all" in args.name:
                        if args.apip == 'yes':
                            api_analysis_files(entry.path, api_data, api_names, plot)
                        analysis_files(entry.path, plots_data, plot)
                    else:
                        if os.path.basename(entry) in args.name:
                            if args.apip == 'yes':
                                api_analysis_files(entry.path, api_data, api_names, plot)
                            analysis_files(entry.path, plots_data, plot)
                except:
                    print(entry.path + ':Error can not analysis data, please check the .json format...')
        benchmarks_plots[plot] = plots_data
        api_plots[plot] = api_data
        if (plot == 'cpu_time' or plot == 'real_time') and ('png' in args.select):
            # date to plot
            for a in benchmarks_plots[plot]:
                y_raw_val = benchmarks_plots[plot][a]
                x_raw_val  = range(0, len(numpy.arange(0, len(y_raw_val), 1)))
                plot_2d(x_raw_val, y_raw_val, plot, args.output, name_cut(a))
    
    if args.apip == 'yes':
        if 'png' in args.select:
            # date to plot
            for b in api_names:
                if 'cpu_time' in args.plot:
                    y_raw_val = api_plots['cpu_time'][b]
                    if api_plots['cpu_time'].get(b + '/p') is None:
                        x_raw_val = None
                    else:
                        x_raw_val = api_plots['cpu_time'][b + '/p']
                        print(x_raw_val)
                    plot_2d(x_raw_val, y_raw_val, 'cpu_time', args.output, b)

                if 'real_time' in args.plot:
                    y_raw_val = api_plots['real_time'][b]
                    if api_plots['real_time'].get(b + '/p') is None:
                        x_raw_val = None
                    else:
                        x_raw_val = api_plots['real_time'][b + '/p']
                        print(x_raw_val)
                    plot_2d(x_raw_val, y_raw_val, 'real_time', args.output, b)

    # # get list data
    # print(pandas.json_normalize(benchmarks_plots))

    # get data from all json to excel/csv.
    for entry in files:
        if entry.path.endswith('.json') and entry.is_file():
            # all .json files analysis.
            if "all" in args.name:
                benchmark_data, context_data = analysis_file(entry.path, args.mode)
            else:
                # name jsonfile analysis.
                if os.path.basename(entry) in args.name:
                    benchmark_data, context_data = analysis_file(entry.path, args.mode)
            if (benchmark_data is not None) and (context_data is not None):
                # data into excel
                excelpath = os.path.join(args.output, 'excel/' + os.path.basename(entry).split('.')[0] + '.xlsx')
                is_dir(excelpath)
                with pandas.ExcelWriter(excelpath) as excelfd:
                    context_data.to_excel(excelfd, sheet_name = 'context')
                    benchmark_data.to_excel(excelfd, sheet_name = 'benchmarks', index=False)
                # fix excel row width
                auto_width(excelpath)
                csvpath = os.path.join(args.output, 'csv/' + os.path.basename(entry).split('.')[0] + '.csv')
                is_dir(csvpath)
                # mode='a+': append to existing data
                context_data.to_csv(csvpath, index = False, encoding ='utf-8')
                benchmark_data.to_csv(csvpath, index = False, encoding ='utf-8', mode = 'a+', header = True)
    
        
if __name__ == '__main__':
    main()