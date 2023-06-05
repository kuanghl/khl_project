import json
import os
from openpyxl import load_workbook

txt_url = r"./data3.txt"
sava_url = r'./data3.xlsx'
 
 
def conversion_json(excel_path, sheets_names: list = None, storage_filepath=None):
    """
    :param excel_path: excel文件地址
    :param sheets_names: 列表传入指定sheet页 不传默认所有
    :param storage_filepath: 保存json文件地址 不传默认当前文件夹
    :return:
    """
    wb = load_workbook(excel_path)
    if sheets_names is None:
        sheets_names = wb.sheetnames
    for sheet_name in sheets_names:
        data = wb[sheet_name]
        temp_list = []
        for i in range(1, data.max_row):
            temp_dict = {}
            for j in range(1, data.max_column + 1):
                title_data = data.cell(row=1, column=j).value
                value_data = data.cell(row=i + 1, column=j).value
                if title_data is not None and value_data is not None:
                    temp_dict[title_data] = value_data
                else:
                    continue
            temp_list.append(temp_dict)
        json_data = json.dumps(temp_list, indent=4, ensure_ascii=False)
        if storage_filepath is None:
            storage_filepath = os.path.dirname(__file__)
        with open(storage_filepath + r"./data3.json".format(sheet_name), "w+", encoding="utf-8") as f:
            f.write(json_data)
 
 
if __name__ == "__main__":
 
    conversion_json("./data3.xlsx")
    # conversion_json("/Users/zhangshuai/PycharmProjects/stock_index/case/xxx.xlsx", [sheet1, sheet2], filepath)